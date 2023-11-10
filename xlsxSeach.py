import openpyxl, os
from tkinter import Tk, filedialog

def limpar_console():
    os.system('cls' if os.name == 'nt' else 'clear')

def obter_caminho_arquivo():
    root = Tk()
    root.withdraw()  # Oculta a janela principal

    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione um arquivo .xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
    )

    return caminho_arquivo

def obter_caminho_salvar_arquivo():
    root = Tk()
    root.withdraw()  # Oculta a janela principal

    caminho_arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
        title="Escolha um local para salvar o arquivo .xlsx"
    )

    return caminho_arquivo

def converter_coluna_para_indice(coluna):
    """Converte o nome da coluna para um índice baseado em letras."""
    indice = 0
    for letra in coluna:
        indice = indice * 26 + (ord(letra.upper()) - ord('A') + 1)
    return indice

def ler_arquivo_excel(nome_arquivo, colunas):
    workbook = openpyxl.load_workbook(nome_arquivo)
    sheet = workbook.active

    dados_matriz = []

    for coluna in colunas:
        indice_coluna = converter_coluna_para_indice(coluna)
        dados_coluna = [sheet.cell(row=i, column=indice_coluna).value for i in range(1, sheet.max_row + 1)]
        dados_matriz.append(dados_coluna)

    workbook.close()

    return dados_matriz

def escrever_em_arquivo_excel(nome_arquivo_saida, dados_matriz):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for col_idx, coluna in enumerate(dados_matriz, start=1):
        for linha_idx, valor in enumerate(coluna, start=1):
            sheet.cell(row=linha_idx, column=col_idx).value = valor

    workbook.save(nome_arquivo_saida)

def menu():
    print("############### Bem vindo ao Sistema de Raspagem de Arquivos .xlsx ###############")
    print("Version 0.2", end="\n")
    print('\n')
    print("ESCOLHA UMA OPÇÃO:" + "\n1-Escolher arquivo .xlsx" + "\n2-Sair")

    menu_opcao = int(input("Digite sua opção: "))
    if menu_opcao == 1:
        nome_arquivo_entrada = obter_caminho_arquivo()
        colunas_para_ler_str = input("Digite as letras das colunas separadas por espaço (por exemplo, A B C): ")

        # Converter as letras das colunas para uma lista de colunas
        colunas_para_ler = colunas_para_ler_str.split()

        dados_matriz = ler_arquivo_excel(nome_arquivo_entrada, colunas_para_ler)

        # Exibir os dados
        print("Dados da matriz:")
        for linha in dados_matriz:
            print(linha)

        print("\nDeseja Salvar o Arquivo .xlsx Filtrado ? (S/N)")
        opcao_save = input()

        if opcao_save.lower() == "s":
            # Obter o nome do arquivo de saída
            nome_arquivo_saida = obter_caminho_salvar_arquivo()
            if not nome_arquivo_saida:
                print("Operação cancelada. Sistema Encerrando ...")
                exit()
            # Escrever os dados filtrados no novo arquivo
            escrever_em_arquivo_excel(nome_arquivo_saida, dados_matriz)

            print(f"Os dados filtrados foram escritos no arquivo {nome_arquivo_saida}.\n")
            limpar_console()
            print('\n')
            menu()
        elif opcao_save.lower() == "n":
            limpar_console()
            print('\n')
            menu()
    else:
        print("Sistema Encerrando ...")
        exit()
menu()